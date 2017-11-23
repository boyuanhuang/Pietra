# -*- coding: utf-8 -*-

import openpyxl, time,os, math
#import string
from openpyxl import Workbook
from openpyxl.compat import range
#from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from indexation import get_txt_important_word
import shutil


#os.mkdir(r'/home/bo/桌面/test/Index')

#print(get_txt_important_word(r'/home/bo/桌面/test/final test/5. 2016.03.07 LRP Mentions Légales et Données Personnelles.txt'))

#string = 'abc'
#print(string[0])
def split_by_headword_alphabet(wordlist):
    global atoz
    n = len(wordlist)
    result = []
    anythingelse = []
    headword = 'a'
    headword_catagory = []
    i = 0
    while wordlist[i][0] < headword:
        anythingelse.append(wordlist[i])
        i+=1
        if i == n :
                break
    while i < n:
        while headword == wordlist[i][0] :
            headword_catagory.append(wordlist[i])
            i+=1
            if i == n :
                break
        if len(headword_catagory)>0:
            result.append(headword_catagory)
        if i == n:
            break
        headword = wordlist[i][0]        
        headword_catagory = []
        while not headword in atoz:
            anythingelse.append(wordlist[i])
            i+=1
            if i == n :
                break
            headword = wordlist[i][0]
            
    if len(anythingelse)>0:
        result.append(anythingelse)
    return result

def split_by_secondword_alphabet(wordlist):
    global atoz
    n = len(wordlist)
    result = []
    anythingelse = []
    headword = 'a'
    headword_catagory = []
    i = 0
    while wordlist[i][1] < headword:
        anythingelse.append(wordlist[i])
        i+=1
        if i == n :
                break
    while i < n:
        while headword == wordlist[i][1] :
            headword_catagory.append(wordlist[i])
            i+=1
            if i == n :
                break
        if len(headword_catagory)>0:
            result.append(headword_catagory)
        if i == n:
            break
        headword = wordlist[i][1]        
        headword_catagory = []
        while not headword in atoz:
            anythingelse.append(wordlist[i])
            i+=1
            if i == n :
                break
            headword = wordlist[i][1]
    if len(anythingelse)>0:
        result.append(anythingelse)
    return result

def split_by_head_then_second(wordlist):
    l = split_by_headword_alphabet(wordlist)
    for i in range(len(l)):
        if l[i][0][0] in atoz:
            l[i] = split_by_secondword_alphabet(l[i])
    return l
    
    
def creat_blank_Excel(name, motherdir):
    dest_filename = os.path.join(motherdir, name+'.xlsx')
    if os.path.exists(dest_filename):
        return
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "indexation word-path"
    ws2 = wb.create_sheet(title="length recorder")
    ws2["A1"].value = 0
    ws3 = wb.create_sheet(title="temporaire")
    wb.save(filename = dest_filename)


atoz= ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z', 'à', 'â', 'æ', 'ç', 'è', 'é', 'ê', 'ë', 'î', 'ï', 'ô', 'ù', 'û', 'ü', 'ÿ', 'œ']

atoz2 = []
for i in atoz:
    for j in atoz:
        atoz2.append(i+j)
'''
file = open(r'/home/bo/桌面/test/testatozplus1', 'a')

for word in atoz2:
    file.write(word + '\n')
file.close()
'''
#l =get_txt_important_word(r'/home/bo/桌面/test/final test/DOC/Uzik_CGV_avec_cessio.txt')
#print(split_by_head_then_second(l))

def find_word_position_in_Excel(word, nb_mots_actuel, ws1): 
    if nb_mots_actuel ==0:
        return 0.5
    rang = [1, nb_mots_actuel]
    while not (rang[1]-rang[0]) < 2:
        middle = math.floor((rang[1]+rang[0])/2)
        if word < str(ws1["A"+str(middle)].value):
            #print(str(ws1["A"+str(middle)].value))
            rang[1] = middle
            #print(rang)
        elif word > str(ws1["A"+str(middle)].value):
            #print(str(ws1["A"+str(middle)].value))
            rang[0] = middle
            #print(rang)
        else: 
            return middle
    if word < str(ws1["A"+str(rang[0])].value ):
        return rang[0]-0.5
    elif word == str(ws1["A"+str(rang[0])].value ):
        return rang[0]
    elif word > str(ws1["A"+str(rang[0])].value) and word < str(ws1["A"+str(rang[1])].value):
        return (rang[1]+rang[0])/2
    elif word == str(ws1["A"+str(rang[1])].value) :
        return rang[1]
    else :
        return rang[1]+0.5
   
    

def indexer_to_headword(headword, thislist, dir_to_write, filepath):  
    Excelpath = os.path.join(dir_to_write, headword+'.xlsx')
    if not os.path.exists(Excelpath):
        creat_blank_Excel(headword, dir_to_write)
        
    wb = load_workbook(filename = Excelpath)
    ws1 = wb["indexation word-path"]
    ws2 = wb["length recorder"]
    nb_mots_actuel = int( ws2["A1"].value)
    
    for word in thislist:
        position_to_add = find_word_position_in_Excel(word, nb_mots_actuel, ws1) #replace it with a local function
        if type(position_to_add) == float: #insert new row
            position_to_add += 0.5
            wordtoinsert = word
            pathtoadd = filepath
            for i in range(int(position_to_add), nb_mots_actuel+2):
                temp1 = ws1["A"+str(i)].value
                temp2 = ws1["B"+str(i)].value
                ws1.cell(column=1, row=i, value= wordtoinsert)
                ws1.cell(column=2, row=i, value= pathtoadd)   
                wordtoinsert = temp1
                pathtoadd = temp2
            nb_mots_actuel+=1
            ws2["A1"].value = nb_mots_actuel
        else: 
            ws1.cell(column = 2, row = position_to_add).value = ws1["B"+str(position_to_add)].value + ", " + filepath            
            
    wb.save(filename = Excelpath) 
    return



def indexer_a_file(filepath, root_index_file_path):
    wordlist = get_txt_important_word(filepath)
    if wordlist ==[]:
        return
    wordlist = split_by_head_then_second(wordlist)
    #print(wordlist)
    n = len(wordlist)
    for i in range(n-1): 
        m = len(wordlist[i])
        for j in range(m-1):
            thislist = wordlist[i][j]
            headword = thislist[0][:2]
            dir_to_write = os.path.join(root_index_file_path,thislist[0][0])
            if not os.path.exists(dir_to_write):
                os.mkdir(dir_to_write)
            indexer_to_headword(headword, thislist, dir_to_write, filepath)
            
        thislist = wordlist[i][m-1]
        headword = thislist[0][:2]
        if headword in atoz2:
            dir_to_write = os.path.join(root_index_file_path, thislist[0][0])
            if not os.path.exists(dir_to_write):
                os.mkdir(dir_to_write)
            indexer_to_headword(headword, thislist, dir_to_write, filepath)
        else:
            print(headword)
            indexer_to_headword('0', thislist, root_index_file_path, filepath)
            
    if type(wordlist[n-1][0]) == str:
        indexer_to_headword('0', wordlist[n-1], root_index_file_path, filepath)     
    else:
        for thislist in wordlist[n-1]:
            headword = thislist[0][:2]
            dir_to_write = os.path.join(root_index_file_path, thislist[0][0])
            if not os.path.exists(dir_to_write):
                os.mkdir(dir_to_write)
            indexer_to_headword(headword, thislist, dir_to_write, filepath)


def indexer_un_folder(folderpath,root_index_file_path):
    filename = os.listdir(folderpath)
    #i =0
    pathlist = []
    for file in filename:
        pathlist.append(os.path.join(folderpath, file))
    for path in pathlist:
        if os.path.isfile(path):
            #i+=1
            #print(i)
            indexer_a_file(path, root_index_file_path)
        elif os.path.isdir(path):
            indexer_un_folder(path, root_index_file_path)

def create_index_for_folder(folderpath, destination):
    name = os.path.basename(folderpath)
    root_index_file_path = os.path.join(destination, name)
    if os.path.exists(root_index_file_path):
        shutil.rmtree(root_index_file_path)
    os.mkdir(root_index_file_path)
    indexer_un_folder(folderpath, root_index_file_path)
    
#os.mkdir(r'/home/bo/桌面/test/Index_GOOGLE')
pbpath = r'/home/bo/桌面/test/GOOGLE/Google pour les pros/1. Ateliers GPLP/Atelier GPLP - Marseille/Signalétique salle/Signalétique sur table/cadre_plexi_def.txt'
nopbpath = r'/home/bo/桌面/test/final test/5. 2016.03.07 LRP Legal Notice and Data Privacy.txt'
path2 = r'/home/bo/桌面/test/final test/2015_THE_MAKEUP_BOOK.txt'
path3 = r'/home/bo/桌面/test/final test/In the sky Paris apartment. 2 bdrm. à Paris.txt'
path4 = r'/home/bo/桌面/test/testatozplus1'

t=time.time()
#l = get_txt_important_word(path4)
#print(l)
#print(split_by_head_then_second(l))
create_index_for_folder(r'/home/bo/桌面/test/GOOGLE/Google pour les pros', r'/home/bo/桌面/test/INDEX')
#indexer_un_folder(r'/home/bo/桌面/test/final test')
#indexer_un_folder(r'/home/bo/桌面/test/GOOGLE/Google pour les pros/1. Ateliers GPLP')
print(time.time()-t)
    

    
"""CODE GRAVE """################################################################################    

''' 
def indexer_un_folder_par_premier(folderpath):
    filename = os.listdir(folderpath)
    i =0
    pathlist = []
    for file in filename:
        pathlist.append(os.path.join(folderpath, file))
    print(pathlist)
    for path in pathlist:
        if os.path.isfile(path):
            i+=1
            print(i)
            try:
                indexer_par_premier_alphbt(path, r'/home/bo/桌面/test/Index')
            except:
                pass
        elif os.path.isdir(path):
            indexer_un_folder_par_premier(path)
            
def indexer_par_premier_alphbt(filepath, Index_root_dir): # Index_root_dir = r'/home/bo/桌面/test/Index'
    list_split_by_alphbt = split_by_headword_alphabet(get_txt_important_word(filepath)) 
    for thislist in list_split_by_alphbt:
        if thislist[0][0] in atoz:
            indexer_to_headword(thislist[0][0], thislist, Index_root_dir, filepath)
        else:
            indexer_to_headword('0', thislist, Index_root_dir, filepath)
    return
'''   
    