# -*- coding: utf-8 -*-




import textract, time, os,re

import openpyxl, math
import io

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

def new_filter(words):       
    mots_to_remove = []  
    mots_to_keep   = []
    for i in range(len(words)):
        if not len(words[i]) in range(3,22) or 'http' in words[i] or ".com" in words[i] or "@" in words[i] or words[i][2]==":" :
            mots_to_remove.append(words[i])
        try:
            if words[i][1] in ["’", "'"]:
                words[i] = words[i][2:]

            elif words[i][2] in  ["’", "'"] :
                words[i] = words[i][3:]
                
            elif words[i][-2] in [ "'", "’"]:
                words[i] = words[i][:-2]
                
            elif words[i][-3] in [ "'", "’"]:
                words[i] = words[i][:-3]  
        except:
            pass
        try:
            nb = int(words[i])      
            if nb in range(1800, 2050) :           # Garder les annees
                mots_to_keep.append(words[i])
        except:
            pass
        
        if not len({'0','1','2','3','4','5','6','7','8','9'}.intersection(list(words[i]))) ==0:
            mots_to_remove.append(words[i])
    
    for mots in mots_to_remove:
        try:
            words.remove(mots)    
        except:
            pass
        
    for mots in mots_to_keep:
        try:
            words.append(mots)    
        except:
            pass 
        
    return words   
    
####################################################################################

    

def get_txt_important_word(filepath):
    global verbe, Fr_uselesswords, En_uselesswords
    break_by_line = []
    try:
        break_by_line = textract.process(filepath, extension='txt').decode("utf-8").lower().split('\n') #it's a list
    except:
        pass
    words = []
    for i in range(len(break_by_line)):
        words+= break_by_line[i].split(' ')
    words = new_filter(words)
    
    for i in range(len(words)):
        words[i] = re.sub(r'\W+', '', words[i])
    words = set(words)
    words = words - Fr_uselesswords - verbe - En_uselesswords
    words = new_filter(list(words))
    return sorted(words)

   
# PENSER A FAIRE LE FILTRE ANGLAIS !!!!
verbe = {'pourra', 'pourrez', 'pourrions','pourront', 'pouvant', 'pouvons','allez', 'allons', 'arrive', 'avez', 'avons', 'dit', 'doit', 'faisons', 'fait', 'faite', 'fasse', 'faut', 'met', 'mis', 'ont', 'peut', 'peuvent', 'pouvez', 'prend', 'sait', 'savent', 'soit', 'veulent', 'veut', 'vient', 'voir', 'voit', 'vont'}

Fr_uselesswords = {'', 'ainsi', 'afin', 'autre', 'alors', 'après', 'arriver', 'assez', 'attendu', 'aucun', 'aucune', 'aussi', 'autres', 'aux', 'avant', 'avec', 'avoir', 'beaucoup', 'ben', 'bien', 'bon', "c'est", 'cent', 'ces', 'cette', 'chaque', 'chez', 'chose', 'cinq', 'circa', 'comme', 'comment', 'concernant', 'contre', 'c’est', "d'ailleurs", 'dans', 'dedans', 'dehors', 'delà', 'depuis', 'derrière', 'des', 'dessous', 'dessus', 'deux', 'devant', 'devers', 'devoir', 'dire', 'dix', 'dixit', 'donc', 'durant', 'dès', 'déjà', 'elle', 'elles', 'emmi', 'encore', 'enfin', 'entenant', 'entre', 'entremi', 'envers', 'environ', 'est', 'excepté', 'faire', 'falloir', 'fors', 'hormis', 'hors', 'ils', 'joignant', 'jouxte', 'jusque', 'les', 'leur', 'leurs', 'lui', 'là-bas', 'mais', 'malgré', 'mettre', 'moins', 'mois', 'mon', 'mots', 'moyennant', 'même', 'niveau', 'non', 'nonobstant', 'nos', 'notre', 'nous', 'oui', 'outre', 'par', 'parce', 'parmi', 'partir', 'pas', 'passé', 'pendant', 'peu', 'peut-être', 'plein', 'plus', 'pour', 'pouvoir', 'prendre', 'puis', "qu'est-ce", 'quand', 'quatre', 'que', 'quel', 'quelle', 'quelles', 'quelque', 'quels', 'question', 'qui', 'quoi', 'qu’un', 'qu’une', 'rester', 'rien', 'sans', 'sauf', 'selon', 'ses', 'soir', 'son', 'sont', 'sous', 'sub', 'suivant', 'sur', 'tandis', 'tel', 'telles', 'tels', 'temps', 'tenir', 'toujours', 'tous', 'tout', 'toute', 'toutes', 'trois', 'trop', 'trouver', 'très', 'un', 'une', 'venir', 'vers', 'versus', 'via', 'vingt', 'voici', 'voilà', 'vos', 'votre', 'vous', 'vraiment', 'ème', 'être', 'ôté'}


En_uselesswords = {'able', 'about', 'above', 'account', 'across', 'addition', 'after', 'against', 'all', 'along', 'also', 'among', 'and', 'any', 'anyway', 'are', 'around', 'at', 'because', 'before', 'behind', 'below', 'beneath', 'beside', 'besides', 'between', 'beyond', 'but', 'by', 'call', 'can', 'case', 'come', 'despite', 'down', 'due', 'during', 'early', 'except', 'exception', 'fact', 'few', 'find', 'for', 'from', 'front', 'get', 'give', 'good', 'has', 'have', 'her', 'here', 'high', 'his', 'how', 'in', 'inside', 'instead', 'into', "isn't", 'its', 'like', 'near', 'nor', 'not', 'of', 'off', 'on', 'one', 'other', 'others', 'our', 'out', 'outside', 'over', 'own', 'past', 'place', 'same', 'say', 'see', 'seem', 'she', 'since', 'small', 'spite', 'still', 'take', 'tell', 'than', 'that', 'the', 'their', 'them', 'there', 'these', 'they', 'thing', 'this', 'three', 'through', 'throughout', 'to', 'too', 'toward', 'try', 'two', 'under', 'underneath', 'until', 'up', 'use', 'very', 'want', 'was', 'were', 'what', 'when', 'where', 'which', 'while', 'who', 'why', 'will', 'with', 'within', 'without', 'would', 'yes', 'you', 'your', 'you’re', 'you’ve'}



pathgoogle = r'/home/bo/桌面/test/GOOGLE/Digital Active'

pbfile = r'/home/bo/桌面/test/GOOGLE/Digital Active/3 - Com Univs/Templates/Template Affiches 4 Couleurs Formation online/Lille/Lille-vert.txt'


path = r'/home/bo/桌面/test/showme4.txt'



dest_Excel_file = r'/home/bo/桌面/test/indexation1.xlsx'

def sort_Excel(dest_Excel_file):
    wb = load_workbook(filename = dest_Excel_file)
    ws1 = wb["indexation word-path"]
    ws2 = wb["length recorder"]
    
    nb_mots_actuel = int( ws2["A1"].value)
    ws3 = wb["temporaire"]
    unsorted_wordlist = [ str(ws1["A"+str(i)].value) for i in range(1, nb_mots_actuel+1)]
    order = [i for i in range(1, nb_mots_actuel+1)]
    
    temp = sorted(zip(unsorted_wordlist, order))
    sorted_word_list = [ temp[i][0]  for i in range(nb_mots_actuel)]
    order = [ temp[i][1]  for i in range(nb_mots_actuel)]
    for i in range(nb_mots_actuel):
         ws3.cell(column=1, row=i+1, value=sorted_word_list[i])
         ws3.cell(column=2, row=i+1 , value= ws1["B"+str(order[i])].value)
    ws1.title = "t"
    ws3.title = "indexation word-path"
    wb.remove_sheet(ws1)
    wb.create_sheet(title="temporaire")

    wb.save(filename = dest_Excel_file)

def find_word_in_Excel(word, dest_Excel_file):  # n = len(sortedwordlist)
    wb = load_workbook(filename = dest_Excel_file)
    ws1 = wb["indexation word-path"]
    ws2 = wb["length recorder"]
    ws3 = wb["temporaire"]
    nb_mots_actuel = int( ws2["A1"].value)
    if nb_mots_actuel ==0:
        return 0.5
    rang = [1, nb_mots_actuel]
    while not (rang[1]-rang[0]) < 2:
        middle = math.floor((rang[1]+rang[0])/2)
        if word < str(ws1["A"+str(middle)].value):
            #print(str(ws1["A"+str(middle)].value))
            rang[1] = middle
            #print(rang)
        elif word > str(ws1["A"+str(middle)].value):
            #print(str(ws1["A"+str(middle)].value))
            rang[0] = middle
            #print(rang)
        else: 
            return middle
    if word < str(ws1["A"+str(rang[0])].value ):
        return rang[0]-0.5
    elif word == str(ws1["A"+str(rang[0])].value ):
        return rang[0]
    elif word > str(ws1["A"+str(rang[0])].value) and word < str(ws1["A"+str(rang[1])].value):
        return (rang[1]+rang[0])/2
    elif word == str(ws1["A"+str(rang[1])].value) :
        return rang[1]
    else :
        return rang[1]+0.5
   

def indexer_by_simple_add(filepath, dest_Excel_file):
    l=[]
    try:
        l = get_txt_important_word(filepath)
    except:
        print(filepath)
        return
    wb = load_workbook(filename = dest_Excel_file)
    ws1 = wb["indexation word-path"]
    ws2 = wb["length recorder"]
    nb_mots_actuel = int( ws2["A1"].value)
    nb_mots_new_added = 0
    for i in range(len(l)):
        word = l[i]
        position_to_add = find_word_in_Excel(word, dest_Excel_file)
        if position_to_add == -1:
            nb_mots_new_added +=1
            position_to_add = nb_mots_actuel+ nb_mots_new_added
            ws1["A"+str(position_to_add)].value = word
            ws1["B"+str(position_to_add)].value = filepath
            
        else:
            ws1.cell(column = 2 , row = position_to_add).value = ws1.cell(column = 2 , row = position_to_add).value + filepath            


    ws2["A1"].value = nb_mots_actuel + nb_mots_new_added
    wb.save(filename = dest_Excel_file)


def indexation_pour_dossier(dossier_path, dest_Excel_file):
    listpath=[]
    timelist = []
    for doc in os.listdir(dossier_path):
        if os.path.isfile(os.path.join(dossier_path,doc)):
            listpath.append(os.path.join(dossier_path,doc))

    for filepath in listpath:
        print(filepath)
        t = time.time()
        indexer_by_simple_add(filepath, dest_Excel_file)
        sort_Excel(dest_Excel_file)
        print(time.time()-t)
        timelist.append(time.time()-t)
    print(timelist)

def search_with_index(word, dest_Excel_file):
    n = find_word_in_Excel(word, dest_Excel_file)
    if n ==-1:
        print( "no result")
    else:
         wb = load_workbook(filename = dest_Excel_file)
         ws1 = wb["indexation word-path"]
         ws2 = wb["length recorder"]
         print("Found")
         lis = []
         for i in range(int(ws2["B"+str(n)].value)):
             lis.append(ws1.cell(column = i+2, row =n).value)
         print(lis)
''' 
t = time.time()
indexer_by_simple_add(r'/home/bo/桌面/test/final test/In the sky Paris apartment. 2 bdrm. à Paris.txt',dest_Excel_file)
print(time.time()-t)
#indexation_pour_dossier(r'/home/bo/桌面/test/final test/DOC', dest_Excel_file)   
'''
'''
listpath=[]
totalword =0
for doc in os.listdir(r'/home/bo/桌面/test/final test/L/PDF à intégrer/FactsheetCB'):
    if os.path.isfile(os.path.join(r'/home/bo/桌面/test/final test/L/PDF à intégrer/FactsheetCB',doc)):
        totalword += len(get_txt_important_word(os.path.join(r'/home/bo/桌面/test/final test/L/PDF à intégrer/FactsheetCB',doc)))
print(totalword)

t = time.time()
indexer_by_simple_add(r'/home/bo/桌面/test/final test/3. 2016.03.10 Contenus Restaurants_FR_relu.txt', dest_Excel_file)
print("mention lagale :")
print(time.time()-t)

#print(get_txt_important_word(r'/home/bo/桌面/test/final test/5. 2016.03.07 Mentions Légales et Données Personnelles.txt'))

t = time.time()
listpath=[]
for doc in os.listdir(r'/home/bo/桌面/test/final test'):
    if os.path.isfile(os.path.join(r'/home/bo/桌面/test/final test',doc)):
        listpath.append(os.path.join(r'/home/bo/桌面/test/final test',doc))
print(listpath)

for filepath in listpath:    
    print(time.time()-t)
    t = time.time()    
    print(get_txt_important_word(filepath))
'''


'''


    
def indexer_par_insertion(filepath, dest_Excel_file):
    l=[]
    try:
        l = get_txt_important_word(filepath)
    except:
        print(filepath)
        return
    wb = load_workbook(filename = dest_Excel_file)
    ws1 = wb["indexation word-path"]
    ws2 = wb["length recorder"]
    nb_mots_actuel = int( ws2["A1"].value)


    for i in range(len(l)):
        word = l[i]
        position_nb = 1
        position = "A"+str(position_nb)
        while True:
            if ws1[position].value == None:
                ws1[position].value = word
                ws1["B"+str(position_nb)].value = filepath
                nb_mots_actuel +=1 
                break
            elif word < ws1[position].value:
                wordtoinsert = word
                pathtoadd = filepath
                for i in range(position_nb, nb_mots_actuel+2):
                    temp1 = ws1["A"+str(i)].value
                    temp2 = ws1["B"+str(i)].value
                    ws1.cell(column=1, row=i, value= wordtoinsert)
                    ws1.cell(column=2, row=i, value= pathtoadd)   
                    wordtoinsert = temp1
                    pathtoadd = temp2
                nb_mots_actuel +=1 
                break
        
            elif word == ws1[position].value : 
                ws1["B"+str(position_nb)].value = ws1["B"+str(position_nb)].value + ", " + filepath
                break
        
            else :
                position_nb+=1
                position = "A"+str(position_nb)
    
    ws2["A1"].value = nb_mots_actuel
    wb.save(filename = dest_Excel_file)
'''